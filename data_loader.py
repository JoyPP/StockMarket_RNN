import os
import string
import pytz
import fasttext
import pandas as pd
import numpy as np
import cPickle
import time
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from openpyxl.reader.excel import load_workbook


def fasttext_model_pretraining():
    '''
    use all extracted text to train fasttext model, saved into model.bin
    :return:
    '''
    directory = 'summary/'
    if os.path.isdir(directory):
        files = [directory+f for f in os.listdir(directory) if f.endswith('.xlsx')]
    else:
        print 'Summary directory Not Found!'
        return
    alltext = 'alltext.txt'
    if os.path.exists(alltext):
        os.remove(alltext)
    translator = string.maketrans(string.punctuation, " " * len(string.punctuation))
    for f in files:
        print 'file is ', f
        wb = load_workbook(filename=f)
        ws = wb.get_active_sheet()
        row, col = ws.max_row, ws.max_column
        col_dict = {1:'A', 2:'B', 3:'C',4:'D', 5:'E', 6:'F', 7: 'G', 8:'H', 9:'I', 10:'J',11:'K',12:'L',13:'M',14:'N'}
        col_range = [1] + range(5, col + 1)
        tmp = ""
        for i in range(2, row+1):
            for j in col_range:
                val = ws[col_dict[j]+str(i)].value
                if val is not None:
                    tmp += val + '\n'
            if i%100 == 0:
                with open(alltext, 'a') as fd:
                    fd.write(tmp.encode('utf-8').translate(translator))
                tmp = ""
        with open(alltext, 'a') as fd:
            fd.write(tmp.encode('utf-8').translate(translator))
    model = fasttext.skipgram(alltext, 'model')
    return model

def time_diff(symbol, bin_size=5, save = False, directory='summary/'):
    '''
    to compute the time differences of two consecusive articles for each company
    :param symbol: stock symbol of specefic company
    :param bin_size: bin size
    :param save: whether save the picture
    :param directory:
    :return:
    '''
    print 'loading data for ', symbol
    summary_file = directory + symbol + '.xlsx'
    wb = load_workbook(filename=summary_file)
    ws = wb.get_active_sheet()
    row, col = ws.max_row, ws.max_column
    time_list = []
    time_diff = []
    for i in range(2, row+1):
        t = ws['C' + str(i)].value.encode('utf-8')
        YY, MM, DD, hh, mm, ss = int(t[:4]), int(t[5:7]), int(t[8:10]), int(t[11:13]), int(t[14:16]), int(t[17:19])
        utc_dt = datetime(YY, MM, DD, hh, mm, ss)
        if len(time_list) > 0:
            time_diff.append((time_list[-1] - utc_dt).total_seconds()/24/3600.0)
        time_list.append(utc_dt)
    bins = np.arange(0, 365, bin_size)  # fixed bin size
    plt.xlim([0, max(time_diff) + bin_size])
    plt.hist(time_diff, bins=bins, alpha=0.5)
    plt.title('Time Interval for ' + symbol)
    plt.xlabel('Days (bin size = %d)'%bin_size)
    plt.ylabel('Count')
    if save:
        save_dir = 'time_diff/'
        if not os.path.exists(save_dir):
            os.mkdir(save_dir)
        plt.savefig(save_dir+symbol + '_time_interval.png')
    plt.show()

def trend(target_price, basic_price, threshold = 0.01):
    '''
    :param target_price: price after time interval
    :param basic_price: price when predicting
    :param threshold: threshold of price changes
    :return: label of trend. 0: stay; 1: down; 2: up
    '''
    percent = (float(target_price) -  float(basic_price)) / float(basic_price)
    if percent >= threshold:
        return 2
    elif percent <= -threshold:
        return 1
    else:
        return 0

def summary_preprocessing(symbol, model, directory = 'dataset/'):
    '''
    load summaries for the symbol
    :param symbol: stock symbol
    :return:
        summary_info: list saving summary, one matrix (seq_len, feature_dim) for each summary
        time_info: list saving time (transfer to US/Eastern timezone,
        author_info saving summary, time, author information respectively
    '''
    print 'Loading data for', symbol
    # read article information from summary_file
    summary_file = directory + symbol + '.xlsx'
    wb = load_workbook(filename=summary_file)
    ws = wb.get_active_sheet()
    row, col = ws.max_row, ws.max_column    # get max row and column number
    col_dict = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L',
                13: 'M', 14: 'N'}
    col_range = [1] + range(5, col + 1) # related column number
    # transfer all punctuations to whitespace
    translator = string.maketrans(string.punctuation, " " * len(string.punctuation))
    # transfer timezone from utc to US/Eastern timezone
    utc = pytz.utc
    eastern = pytz.timezone('US/Eastern')
    fmt = '%Y-%m-%d %H:%M:%S %Z%z'
    # list to save summary vector, US/Eastern time, and author information respectively
    summary_info, time_info, author_info = [], [], []
    # read summary_file
    for i in range(2, row + 1):
        # read summary and save its vector matrix into the summary_info
        summary = ""
        for j in col_range:
            val = ws[col_dict[j]+str(i)].value
            if val is not None:
                summary += val.encode('utf-8').translate(translator) + ' '  # without eliminating punctuation
        words = summary.split()
        summary = [model[w] for w in words]
        summary_info.append(summary)
        # read time and transfer it to US/Eastern Timezone and save it into the time_info
        t = ws['C'+str(i)].value.encode('utf-8')
        YY, MM, DD, hh, mm, ss = int(t[:4]), int(t[5:7]), int(t[8:10]), int(t[11:13]), int(t[14:16]), int(t[17:19])
        utc_dt = datetime(YY, MM, DD, hh, mm, ss, tzinfo=utc)
        est = utc_dt.astimezone(eastern)
        t = est.strftime(fmt)
        # preprocessing time to US/Eastern timezone, only show date
        # if the article is published before market closing, mark it current date
        # if the article is published after market closing, mark it next day
        # (next trading day to be considered)
        if (t.endswith('EDT-0400') and (t[11:19] <= '16:00:00')) or (
            t.endswith('EDT-0500') and (t[11:19] <= '17:00:00')):
            t = t[:10]
        else:
            t = (est + timedelta(days=1)).strftime(fmt)[:10]
        time_info.append(t) # only save the date
        # read author information and save it into the author_info
        author = ws['D'+str(i)].value.encode('utf-8')
        author_info.append(author)
    summary_info.reverse() #  time sequence
    time_info.reverse()
    author_info.reverse()
    return summary_info, time_info, author_info

def matrix2vec(matrix_list):
    '''
    to average matrix to a vector,
    :param matrix_list: list of matrix, each matrix has dimension: (seq_len, feature_dim)
    :return: list of vector with dimension (feature_dim,)
    '''
    vec_list = []
    for matrix in matrix_list:
        matrix = np.array(matrix)
        vec = np.average(matrix, axis=0)
        vec_list.append(vec)
    return vec_list

def pre_padding(summary_info, batch_size = 16):
    '''
    to padding the seq_matrix in summary_info to same seq_len in each batch
    :param summary_info: list of sequence matrix
    :param batch_size: number of sequence in each batch
    :return:
        result: (n_block, batch_size, max_len-in-this-batch, feature_dim)
            list of padded sequence matrix, sequence ranked as their true seq_len descending
        seq_lens: (n_block, seq_len-list-in-this-batch)
            list of seq_len corresponding to seq_matrix in result list
        sort_list: (n_block, sort_list-in-this-batch)
            list of np.array showing the original indexes of sequences
    '''
    n_data = len(summary_info)
    n_batch = n_data - batch_size + 1
    feature_dim = len(summary_info[0][0])
    result = list() # (n_batch, batch_size, max_len, feature_dim)
    seq_lens = list() # (n_batch, batch_size)
    sort_list = list() # (n_batch, batch_size)
    for i in range(n_batch):
        # start = i*batch_size
        # save the seq len of each batch into seq_len
        '''
        if len(seq_lens) == 0:
            seq_len = list() # for each batch, the sequence len list
            for j in range(i, i+batch_size):
                seq_len.append(len(summary_info[j]))
        else:
            seq_len = seq_len[1:]   # remove the first element
            seq_len.append(len(summary_info[i+batch_size-1]))   # add the last ele in the batch
        '''
        seq_len = []
        for j in range(i, i+batch_size):
            seq_len.append(len(summary_info[j]))
        max_len = max(seq_len)
        #
        # save the padding summary into re  sult
        tmp = list()
        for j in range(i, i+batch_size):
            cur_summary = list(summary_info[j])
            cur_summary.extend([[0]*feature_dim]*(max_len - len(cur_summary)))
            tmp.append(cur_summary)
        # get index of sorted
        seq_len = np.array(seq_len)
        sort_idx = np.argsort(seq_len)[::-1]
        #
        tmp = np.array(tmp)
        np.take(tmp, sort_idx, axis=0, out=tmp)
        np.take(seq_len, sort_idx, axis=0, out=seq_len)
        #
        sort_list.append(sort_idx)
        seq_len = list(seq_len)
        seq_lens.append(seq_len)
        result.append(list(tmp))
    return result, seq_lens, sort_list

def fix_len_padding(summary_info):
    max_len = 100
    feature_dim = len(summary_info[0][0])
    for i, s in enumerate(summary_info):
        if len(s) > max_len:
            summary_info[i] = s[:max_len]
        else:
            summary_info[i].extend([[0]*feature_dim]*(max_len - len(s)))
    return summary_info

def price_preprocessing(symbol, time_info, time_interval, directory = 'dataset/'):
    # read stock price information from price_file
    price_file = directory + symbol + '.csv'
    dateparse = lambda dates: pd.datetime.strptime(dates, '%Y-%m-%d')
    price_data = pd.read_csv(price_file, parse_dates=[0], index_col='Date', date_parser=dateparse, usecols=[0,4])
    price_data = price_data['Close']
    # get price after timedelta days from each of the time_info
    fmt = '%Y-%m-%d'
    targets_dict = {}
    time_intervals = set([7, 14, 21] + [time_interval])
    idx = price_data.index
    for t in time_info:
        for delta in time_intervals:
            target_time = (datetime.strptime(t, fmt) + timedelta(days=delta)).strftime(fmt)
            # if market is not open at t, then t-1
            while t not in idx:
                t = (datetime.strptime(t, fmt) - timedelta(days=1)).strftime(fmt)
            # if market is not open at target_time, then target_time+1
            while target_time not in idx:
                target_time = (datetime.strptime(target_time, fmt) + timedelta(days=1)).strftime(fmt)

            # set label
            label = trend(price_data[target_time], price_data[t])

            # save label into targets_dict
            if str(delta) in targets_dict:
                targets_dict[str(delta)].append(label)
            else:
                targets_dict[str(delta)] = list([label])
    return targets_dict

def shuffle_samples(train_samples, train_labels, test_samples, test_labels):
    '''
    type(inputs) == 'list', shuffle then return list
    :param train_samples:
    :param train_labels:
    :param test_samples:
    :param test_labels:
    :return:
    '''
    # transfer to numpy.array
    train_samples = np.array(train_samples, dtype=np.float32)
    train_labels = np.array(train_labels, dtype=np.int32)
    test_samples = np.array(test_samples, dtype=np.float32)
    test_labels = np.array(test_labels, dtype=np.int32)

    train_num = train_samples.shape[0]
    test_num = test_samples.shape[0]
    print train_samples.shape, test_samples.shape

    # shuffling data
    print 'train data shuffling'
    train_idx = np.arange(train_num, dtype=np.int32)
    np.random.shuffle(train_idx)
    np.take(train_samples, train_idx, axis=0, out=train_samples)
    np.take(train_labels, train_idx, axis=0, out=train_labels)
    train_samples = train_samples.tolist()
    train_labels = train_labels.tolist()

    print 'test data shuffling'
    test_idx = np.arange(test_num, dtype=np.int32)
    np.random.shuffle(test_idx)
    np.take(test_samples, test_idx, axis=0, out=test_samples)
    np.take(test_labels, test_idx, axis=0, out=test_labels)
    test_samples = test_samples.tolist()
    test_labels = test_labels.tolist()

    return train_samples, train_labels, test_samples, test_labels


def data_division(data, batch_size, window_size, shuffle=False):
    '''
    data division with no-padding data
    :param data: tuple of (inputs, targets)
    :param batch_size:
    :param window_size:
    :return: training and test dataset (#batch, 2, batch_size, window_size, feature_dim)
    '''
    # inputs: (n_data, feature_dim)
    # targets: (n_data,)
    inputs, targets = data

    xsample, ysample = [], []
    for index in range(len(inputs) - window_size + 1):
        xsample.append(inputs[index: index + window_size])  # (#sample, window_size, feature_dim)
        ysample.append(targets[index + window_size - 1])    # (#sample, )

    n_sample = len(xsample)
    n_batch = n_sample // batch_size
    xsample = xsample[n_sample - n_batch*batch_size:]   # keep the newest data
    ysample = ysample[n_sample - n_batch*batch_size:]

    # divide into training and test dataset
    row = round(0.8 * n_batch)
    train_samples = xsample[:int(row)*batch_size]
    train_labels = ysample[:int(row)*batch_size]

    test_samples = xsample[int(row)*batch_size:]
    test_labels = ysample[int(row)*batch_size:]

    if shuffle:
        train_samples, train_labels, test_samples, test_labels = shuffle_samples(train_samples, train_labels, test_samples, test_labels)

    train_dataset = []
    for b in range(int(row)):
        x = train_samples[b*batch_size: (b+1)*batch_size]    # (#train_batch, batch_size, window_size, feature_dim)
        y = train_labels[b*batch_size: (b+1)*batch_size]   # (#train_batch, batch_size, 1)
        train_dataset.append((x, y))

    test_dataset = []
    for b in range(n_batch-int(row)):
        x = test_samples[b*batch_size: (b+1)*batch_size]    # (test_batch, batch_size, window_size, feature_dim)
        y = test_labels[b*batch_size: (b+1)*batch_size]   # (test_batch, batch_size, 1)
        test_dataset.append((x, y))

    return train_dataset, test_dataset

"""
def padding_data_division(data, seq_lens, sort_list, window_size = 10):
    '''
    data division with padding data
    :param data: tuple of (inputs, targets)
    :param seq_lens: list of seq_len corresponding to seq_matrix in result list
    :param sort_list: list of np.array showing the original indexes of sequences
    :param window_size:
    :return: training and test dataset: (n_batch, window_size, batch_size, max_len, feature_dim)
    '''
    # inputs: (n_block, batch_size, max_len-in-this-batch, feature_dim)
    # targets : (n_data,)   # n_block = n_data - batch_size + 1
    inputs, targets = data

    n_batch, batch_size = len(inputs), len(inputs[0])

    x, y = [], []
    seqlen, sortlist = [], []
    for index in range(0, n_batch - window_size + 1, batch_size):
        x.append(inputs[index: index+window_size])
        y.append(targets[index+window_size-1: index+window_size-1+batch_size])
        seqlen.append(seq_lens[index: index+window_size])
        sortlist.append(sort_list[index: index + window_size])

    # divide into training and test dataset
    # x: (n_batch, window_size, batch_size, padding_seq_len, feature_dim)
    # y: (n_batch, batch_size)
    row = round(0.8 * len(x))
    xtrain = x[:int(row)]
    ytrain = y[:int(row)]
    sqlentrain = seqlen[:int(row)]
    sttrain = sortlist[:int(row)]

    xtest = x[int(row):]
    ytest = y[int(row):]
    sqlentest = seqlen[int(row):]
    sttest = sortlist[int(row):]

    train_dataset = []
    for xtr, sqtr, sttr, ytr in zip(xtrain, sqlentrain, sttrain, ytrain):
        train_dataset.append((xtr, sqtr, sttr, ytr))

    test_dataset = []
    for xtr, sqtr, sttr, ytr in zip(xtest, sqlentest, sttest, ytest):
        test_dataset.append((xtr, sqtr, sttr, ytr))   # label should be list

    return train_dataset, test_dataset
"""

def data_loader_for_each_symbol_ave_seq(symbol, model, directory, batch_size = 4, time_interval = 7, window_size = 10):
    '''
    :param symbol: symbol of the stock
    :param model: fasttext model
    :param directory: directory of file saving summary and price
    :param time_interval: prediction time interval
    :return: train_dataset, test_dataset consisting of tuples of (inputs, targets)
    '''
    # get summary information for each symbol
    summary_info, time_info, author_info = summary_preprocessing(symbol, model, directory)

    # tranfer summary matrix to summary vector
    summary_vec = matrix2vec(summary_info)

    # get label (up or down) for each symbol
    targets_dict = price_preprocessing(symbol, time_info, time_interval, directory)

    # divide data into training and test dataset and return
    return data_division((summary_vec, targets_dict[str(time_interval)]), batch_size, window_size, True)

def data_loader_for_each_symbol_cnn(symbol, model, directory, batch_size = 1, time_interval = 7, window_size = 10):
    '''
    :param symbol: symbol of the stock
    :param model: fasttext model
    :param directory: directory of file saving summary and price
    :param time_interval: prediction time interval
    :return: train_dataset, test_dataset consisting of tuples of (inputs, targets)
    '''
    # get summary information for each symbol
    summary_info, time_info, author_info = summary_preprocessing(symbol, model, directory)

    # padding summary with fix len
    summary_info = fix_len_padding(summary_info)
    print '#data = ', len(summary_info)
    # get label (up or down) for each symbol
    targets_dict = price_preprocessing(symbol, time_info, time_interval, directory)

    # divide data into training and test dataset and return
    return data_division((summary_info, targets_dict[str(time_interval)]), batch_size, window_size)



def data_loader(model, args):
    '''
    return train_dataset, test_dataset, and the path they saved
    :param model: fasttext word2vec model
    :param args:
    :return:
    '''
    pkl_path = 'pklsets_gru' + '_bs%d_ws%d_ti%d/' % (args.batch_size, args.window_size, args.time_interval)
    TrainDatasetFile = pkl_path + 'TrainDataset.pkl'
    TestDatasetFile = pkl_path + 'TestDataset.pkl'
    # if train_dataset and test_dataset exists, load from files
    if os.path.exists(TrainDatasetFile) and os.path.exists(TestDatasetFile):
        print 'loading train and test dataset from pkl files.'
        start = time.time()
        with open(TrainDatasetFile, 'r') as f:
            train_dataset = cPickle.load(f)
        with open(TestDatasetFile, 'r') as f:
            test_dataset = cPickle.load(f)
        print 'loading time = ', time.time() - start
    else:
        print 'collect and prepare the train and test dataset'
        if os.path.exists(args.data_directory) and os.path.isdir(args.data_directory):
            symbols = [f[:-5] for f in os.listdir(args.data_directory) if f.endswith('.xlsx')]
        else:
            print 'wrong data_directory!'
        if not os.path.exists(pkl_path):
            os.mkdir(pkl_path)
        train_dataset, test_dataset = [], []
        for symbol in symbols:
            print 'preparing data for ', symbol
            symbol_train, symbol_test = data_loader_for_each_symbol_cnn(symbol, model, directory=args.data_directory,
                                                batch_size=args.batch_size, time_interval=args.time_interval, window_size=args.window_size)
            train_dataset.extend(symbol_train)
            test_dataset.extend(symbol_test)
        print 'save train and test datasets into pkl files.'
        with open(TrainDatasetFile, 'w') as f:
            cPickle.dump(train_dataset, f)
        with open(TestDatasetFile, 'w') as f:
            cPickle.dump(test_dataset, f)

    return train_dataset, test_dataset, pkl_path
