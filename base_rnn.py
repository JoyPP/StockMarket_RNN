import os
import argparse
import string
import pytz
import pandas as pd
import cPickle
from datetime import datetime, timedelta

import fasttext
from openpyxl.reader.excel import load_workbook


import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import torch.backends.cudnn as cudnn

from torch.autograd import Variable
from utils import progress_bar

def fasttext_model_pretraining():
    '''
    use all extracted text to train fasttext model, saved into model.bin
    :return:
    '''
    directory = 'summary/'
    if os.path.isdir(directory):
        files = [directory+f for f in os.listdir(directory) if f.endswith('.xlsx')]
    alltext = 'alltext.txt'
    if os.path.exists(alltext):
        os.remove(alltext)
    translator = string.maketrans(string.punctuation, " " * len(string.punctuation))
    for f in files:
        print 'file is ', f
        wb = load_workbook(filename=f)
        ws = wb.get_active_sheet()
        row, col = ws.max_row, ws.max_column
        col_dict = {1:'A', 2:'B', 3:'C',4:'D', 5:'E', 6:'F', 7: 'G', 8:'H', 9:'I', 10:'J',11:'K',12:'L',13:'M',14:'N'}
        col_range = [1] + range(5, col + 1)
        tmp = ""
        for i in range(2, row+1):
            for j in col_range:
                val = ws[col_dict[j]+str(i)].value
                if val is not None:
                    tmp += val + '\n'
            if i%100 == 0:
                with open(alltext, 'a') as fd:
                    fd.write(tmp.encode('utf-8').translate(translator))
                tmp = ""
        with open(alltext, 'a') as fd:
            fd.write(tmp.encode('utf-8').translate(translator))
    model = fasttext.skipgram(alltext, 'model')
    return model


def summary_preprocessing(symbol, model, directory = 'datasetq/'):
    '''
    load summaries for the symbol
    :param symbol: stock symbol
    :return: summary_info, time_info, author_info saving summary, time, author information respectively
    '''
    print 'Loading data for', symbol
    # read article information from summary_file
    summary_file = directory + symbol + '.xlsx'
    wb = load_workbook(filename=summary_file)
    ws = wb.get_active_sheet()
    row, col = ws.max_row, ws.max_column    # get max row and column number
    col_dict = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L',
                13: 'M', 14: 'N'}
    col_range = [1] + range(5, col + 1) # related column number
    # transfer all punctuations to whitespace
    translator = string.maketrans(string.punctuation, " " * len(string.punctuation))
    # transfer timezone from utc to US/Eastern timezone
    utc = pytz.utc
    eastern = pytz.timezone('US/Eastern')
    fmt = '%Y-%m-%d %H:%M:%S %Z%z'
    # list to save summary vector, US/Eastern time, and author information respectively
    summary_info, time_info, author_info = [], [], []
    # read summary_file
    for i in range(2, row + 1):
        # read summary and save its vector matrix into the summary_info
        summary = ""
        for j in col_range:
            val = ws[col_dict[j]+str(i)].value
            if val is not None:
                summary += val.encode('utf-8').translate(translator) + ' '  # without eliminating punctuation
        words = summary.split()
        summary = [model[w] for w in words]
        summary_info.append(summary)
        # read time and transfer it to US/Eastern Timezone and save it into the time_info
        t = ws['C'+str(i)].value.encode('utf-8')
        YY, MM, DD, hh, mm, ss = int(t[:4]), int(t[5:7]), int(t[8:10]), int(t[11:13]), int(t[14:16]), int(t[17:19])
        utc_dt = datetime(YY, MM, DD, hh, mm, ss, tzinfo=utc)
        est = utc_dt.astimezone(eastern)
        t = est.strftime(fmt)
        # preprocessing time to US/Eastern timezone, only show date
        # if the article is published before market closing, mark it current date
        # if the article is published after market closing, mark it next day
        # (next trading day to be considered)
        if (t.endswith('EDT-0400') and (t[11:19] <= '16:00:00')) or (
            t.endswith('EDT-0500') and (t[11:19] <= '17:00:00')):
            t = t[:10]
        else:
            t = (est + timedelta(days=1)).strftime(fmt)[:10]
        time_info.append(t) # only save the date
        # read author information and save it into the author_info
        author = ws['D'+str(i)].value.encode('utf-8')
        author_info.append(author)
    summary_info.reverse() #  time sequence
    time_info.reverse()
    author_info.reverse()
    return summary_info, time_info, author_info

#summary_info, time_info, author_info = summary_preprocessing(symbol='BHP',model=model)

def price_preprocessing(symbol, time_info, time_interval, directory = 'datasetq/'):
    # read stock price information from price_file
    price_file = directory + symbol + '.csv'
    dateparse = lambda dates: pd.datetime.strptime(dates, '%Y-%m-%d')
    price_data = pd.read_csv(price_file, parse_dates=[0], index_col='Date', date_parser=dateparse, usecols=[0,4])
    price_data = price_data['Close']
    # get price after timedelta days from each of the time_info
    fmt = '%Y-%m-%d'
    targets_dict = {}
    time_intervals = set([7, 14, 21] + [time_interval])
    idx = price_data.index
    for t in time_info:
        for delta in time_intervals:
            target_time = (datetime.strptime(t, fmt) + timedelta(days=delta)).strftime(fmt)
            # if market is not open at t, then t-1
            while t not in idx:
                t = (datetime.strptime(t, fmt) - timedelta(days=1)).strftime(fmt)
            # if market is not open at target_time, then target_time+1
            while target_time not in idx:
                target_time = (datetime.strptime(target_time, fmt) + timedelta(days=1)).strftime(fmt)
            # set label
            label = 0
            if price_data[target_time] > price_data[t]:
                label = 1
            elif price_data[target_time] < price_data[t]:
                label = -1
            #price = price_data[target_time]
            if str(delta) in targets_dict:
                targets_dict[str(delta)].append(label)
            else:
                targets_dict[str(delta)] = list([label])
    return targets_dict

#targets_dict = price_preprocessing(symbol='BHP', time_info=time_info, time_interval=7)

def data_division(data, window_size):
    inputs, targets = data

    x, y = [], []
    for index in range(len(inputs) - window_size + 1):
        x.append(inputs[index: index+window_size])
        y.append(targets[index+window_size-1])

    # divide into training and test dataset
    row = round(0.9 * len(x))
    xtrain = x[:int(row)]
    ytrain = y[:int(row)]

    xtest = x[int(row):]
    ytest = y[int(row):]

    train_dataset = []
    for xtr, ytr in zip(xtrain, ytrain):
        train_dataset.append((xtr,ytr))

    test_dataset = []
    for xtr, ytr in zip(xtest, ytest):
        test_dataset.append((xtr, ytr))

    return train_dataset, test_dataset

def data_loader(symbol, model, directory, time_interval = 7, window_size = 10):
    '''
    :param symbol: symbol of the stock
    :param model: fasttext model
    :param directory: directory of file saving summary and price
    :param time_interval: prediction time interval
    :return: train_dataset, test_dataset with format (inputs, targets)
    '''
    summary_info, time_info, author_info = summary_preprocessing(symbol, model, directory)
    targets_dict = price_preprocessing(symbol, time_info, time_interval, directory)

    # divide data into training and test dataset and return
    return data_division((summary_info, targets_dict[str(time_interval)]), window_size)


def train(net, use_cuda=False):
    net.train()
    train_loss = 0
    total = 0
    correct = 0
    for batch_idx, (inputs, targets) in enumerate(train_dataset):
        if use_cuda:
            inputs, targets = inputs.cuda(), targets.cuda()
        optimizer.zero_grad()
        targets = Variable(torch.Tensor(targets))
        outputs = net(inputs)
        loss = criterion(outputs, targets)
        loss.backward()
        optimizer.step()

        train_loss += loss.data[0]
        _, predicted = torch.max(outputs.data, 1)

        total += targets.size(0)
        correct += predicted.eq(targets.data).cpu().sum()

        progress_bar(batch_idx, len(train_dataset), 'Loss: %.3f | Acc: %.3f%% (%d/%d)'
                     % (train_loss/(batch_idx+1), 100.*correct/total, correct, total))

def test(net, test_dataset, use_cuda=False):
    net.eval()
    test_loss = 0
    correct = 0
    total = 0
    for batch_idx, (inputs, targets) in enumerate(test_dataset):
        if use_cuda:
            inputs, targets = inputs.cuda(), targets.cuda()
        optimizer.zero_grad()
        targets = Variable(torch.Tensor(targets))
        outputs = net(inputs)
        loss = criterion(outputs, targets)

        test_loss += loss.data[0]
        _, predicted = torch.max(outputs.data, 1)
        total += targets.size(0)
        correct += predicted.eq(targets.data).cpu().sum()

        progress_bar(batch_idx, len(test_dataset), 'Loss: %.3f | Acc: %.3f%% (%d/%d)'
                     % (test_loss/(batch_idx+1), 100.*correct/total, correct, total))



class LSTMModel(nn.Module):
    def __init__(self, input_size, hidden_size, output_size):
        super(LSTMModel, self).__init__()
        self.input_size = input_size
        self.hidden_size = hidden_size

        self.hidden = self.init_hidden()
        self.lstm = nn.LSTM(input_size, hidden_size)
        self.linear = nn.Linear(hidden_size, output_size)
        self.softmax = nn.LogSoftmax()

    def forward(self, sequences):
        # sequence: (seq_number, seq_len, input_size)
        for seq in sequences:
            seq_len = len(seq)
            out, self.hidden = self.lstm(Variable(torch.Tensor(seq).view(seq_len, 1, -1)), self.hidden)
        output = self.softmax(self.linear(self.hidden))
        return output

    def init_hidden(self):
        return (Variable(torch.zeros(1, 1, self.hidden_size)),
                Variable(torch.zeros(1, 1, self.hidden_size)))


if __name__ == '__main__':
    '''
    parser = argparse.ArgumentParser(description='PyTorch CIFAR10 Training')
    parser.add_argument('--max_epoch', default=0, type=int, help='max_epoch')
    parser.add_argument('--lr', default=0.1, type=float, help='learning rate')
    parser.add_argument('--momentum', default=0.9, type=float, metavar='M',
                        help='momentum')
    args = parser.parse_args()

    use_cuda = torch.cuda.is_available()
    '''

    use_cuda = False

    # get fastText model
    if not os.path.exists('model.bin'):
        model = fasttext_model_pretraining()
    else:
        model = fasttext.load_model('model.bin')

    # Load data
    symbol = 'BHP'
    global train_dataset, test_dataset
    if os.path.exists(symbol+'_train.pkl'):
        print 'read datasets from pkl files.'
        with open(symbol+'_train.pkl', 'r') as f:
            train_dataset = cPickle.load(f)
        with open(symbol + '_test.pkl', 'r') as f:
            test_dataset = cPickle.load(f)
    else:
        train_dataset, test_dataset = data_loader('BHP', model, 'datasetq/')
        print 'saving dataset into pkl files.'
        with open(symbol+'_train.pkl', 'w') as f:
            cPickle.dump(train_dataset, f)
        with open(symbol+'_test.pkl', 'w') as f:
            cPickle.dump(test_dataset, f)


    # Check the save_dir exists or not
    # if not os.path.exists(args.save_dir):
    #    os.makedirs(args.save_dir)

    # build model
    print 'Building model...'
    net = LSTMModel(100,200,80)

    if use_cuda:
        net.cuda()
        net = torch.nn.DataParallel(net, device_ids=range(torch.cuda.device_count()))
        cudnn.benchmark = True

    criterion = nn.CrossEntropyLoss()
    #optimizer = optim.SGD(net.parameters(), lr=args.lr, momentum=args.momentum, weight_decay=5e-4)
    optimizer = optim.SGD(net.parameters(), lr=0.1, momentum=0.9, weight_decay=5e-4)

    for epoch in range(0, 100):
        train(net, epoch)
        test(net, epoch)

    print 'Finish...'

































